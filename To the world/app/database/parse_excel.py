import openpyxl
import json
from datetime import datetime

def parse_date(val):
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if val is None:
        return None
    val_str = str(val).strip().replace(",", " ")
    for fmt in ("%d %b %Y", "%d %b  %Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(val_str, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass
    return None

def clean_float(val):
    if val is None:
        return None
    try:
        if isinstance(val, str):
            val = val.strip().replace(",", ".")
        return float(val)
    except ValueError:
        return None

# Load Excel Files
hydrology_file = "./Data for Database /ESWATINI IIMA CROSS BORDER Hydrology Data Dec 2025 - May 2026.xlsx"
dams_file = "./Data for Database /Eswatini Major Dams Data April 2025 - May 2026.xlsx"

wb_hyd = openpyxl.load_workbook(hydrology_file, data_only=True)
wb_dam = openpyxl.load_workbook(dams_file, data_only=True)

data = {
    "stations": [],
    "eflow_requirements": [],
    "measurements": []
}

# 1. Define Stations to Pre-create if missing
# Dams in Eswatini
data["stations"].extend([
    {
        "code": "MNJOLI-DAM-01",
        "name": "Mnjoli Dam",
        "latitude": -26.1667,
        "longitude": 31.6667,
        "category": "dam",
        "water_source": "surface",
        "realtime": False,
        "org": "DWA-SW",
        "country": "Eswatini",
        "basin": "Maputo",
        "subcatchment": "MAP-USUTHU",
        "capabilities": ["dam_level"]
    },
    {
        "code": "LUBOVANE-DAM-01",
        "name": "Lubovane Dam",
        "latitude": -26.7420,
        "longitude": 31.7029,
        "category": "dam",
        "water_source": "surface",
        "realtime": False,
        "org": "DWA-SW",
        "country": "Eswatini",
        "basin": "Maputo",
        "subcatchment": "MAP-USUTHU",
        "capabilities": ["dam_level"]
    },
    {
        "code": "LAVUMISA-DAM-01",
        "name": "Lavumisa Balancing Dam",
        "latitude": -27.3100,
        "longitude": 31.8920,
        "category": "dam",
        "water_source": "surface",
        "realtime": False,
        "org": "DWA-SW",
        "country": "Eswatini",
        "basin": "Maputo",
        "subcatchment": "MAP-NGWAVUMA",
        "capabilities": ["dam_level"]
    },
    {
        "code": "LUPHOHLO-DAM-01",
        "name": "Luphohlo Dam",
        "latitude": -26.3949,
        "longitude": 31.1009,
        "category": "dam",
        "water_source": "surface",
        "realtime": False,
        "org": "DWA-SW",
        "country": "Eswatini",
        "basin": "Maputo",
        "subcatchment": "MAP-LUSUSHWANA",
        "capabilities": ["dam_level"]
    },
    # Missing flow stations in Eswatini
    {
        "code": "GS-08",
        "name": "Nsoko (Ngwavuma)",
        "latitude": -27.0333,
        "longitude": 31.9500,
        "category": "river",
        "water_source": "surface",
        "realtime": False,
        "org": "DWA-SW",
        "country": "Eswatini",
        "basin": "Maputo",
        "subcatchment": "MAP-NGWAVUMA",
        "capabilities": ["flow"]
    },
    {
        "code": "GS-31",
        "name": "Sandlane (Little Usuthu)",
        "latitude": -26.5667,
        "longitude": 30.7833,
        "category": "river",
        "water_source": "surface",
        "realtime": False,
        "org": "DWA-SW",
        "country": "Eswatini",
        "basin": "Maputo",
        "subcatchment": "MAP-USUTHU",
        "capabilities": ["flow"]
    },
    {
        "code": "GS-39",
        "name": "Ndlotane",
        "latitude": -26.1333,
        "longitude": 31.0900,
        "category": "river",
        "water_source": "surface",
        "realtime": False,
        "org": "DWA-SW",
        "country": "Eswatini",
        "basin": "Maputo",
        "subcatchment": "MAP-USUTHU",
        "capabilities": ["flow"]
    },
    {
        "code": "DUMBARTON-01",
        "name": "Dumbarton (Mpuluzi)",
        "latitude": -26.4700,
        "longitude": 30.8200,
        "category": "river",
        "water_source": "surface",
        "realtime": False,
        "org": "DWA-SW",
        "country": "Eswatini",
        "basin": "Maputo",
        "subcatchment": "MAP-MPULUZI",
        "capabilities": ["flow"]
    },
    {
        "code": "NDUMO-01",
        "name": "Ndumo (Pongola)",
        "latitude": -26.8600,
        "longitude": 32.2400,
        "category": "river",
        "water_source": "surface",
        "realtime": False,
        "org": "ARA-Sul",
        "country": "Mozambique",
        "basin": "Maputo",
        "subcatchment": "MAP-PONGOLA",
        "capabilities": ["flow"]
    }
])

# 2. Define Interim Target Instream Flows (from IIMA compliance image)
data["eflow_requirements"] = [
    {
        "river": "Maputo",
        "key_point": "Salamanga (E-4)",
        "station_code": "E-4",
        "mean_annual_mm3": 840.0,
        "min_flow_m3_s": 2.7,
        "subcatchment_code": "MAP-MAPUTO",
        "note": "Interim Target Instream Flow from IIMA Table 5-4."
    },
    {
        "river": "Pongola",
        "key_point": "Ndumo",
        "station_code": "NDUMO-01",
        "mean_annual_mm3": 300.0,
        "min_flow_m3_s": 0.8,
        "subcatchment_code": "MAP-PONGOLA",
        "note": "Interim Target Instream Flow from IIMA Table 5-4."
    },
    {
        "river": "Ngwavuma",
        "key_point": "At the border",
        "station_code": "GS-08",
        "mean_annual_mm3": 50.0,
        "min_flow_m3_s": 0.1,
        "subcatchment_code": "MAP-NGWAVUMA",
        "note": "Interim Target Instream Flow from IIMA Table 5-4."
    },
    {
        "river": "Mkhondvo",
        "key_point": "GS 25",
        "station_code": "GS-25",
        "mean_annual_mm3": 35.0,
        "min_flow_m3_s": 0.1,
        "subcatchment_code": "MAP-MKHONDVO",
        "note": "Interim Target Instream Flow from IIMA Table 5-4."
    },
    {
        "river": "Hlelo",
        "key_point": "GS 22",
        "station_code": "GS-22",
        "mean_annual_mm3": 35.0,
        "min_flow_m3_s": 0.1,
        "subcatchment_code": "MAP-USUTHU",
        "note": "Interim Target Instream Flow from IIMA Table 5-4."
    },
    {
        "river": "Ngwempisi",
        "key_point": "GS 21",
        "station_code": "GS-21",
        "mean_annual_mm3": 30.0,
        "min_flow_m3_s": 0.1,
        "subcatchment_code": "MAP-NGWEMPISI",
        "note": "Interim Target Instream Flow from IIMA Table 5-4."
    },
    {
        "river": "Usuthu",
        "key_point": "GS 23",
        "station_code": "GS-23",
        "mean_annual_mm3": 20.0,
        "min_flow_m3_s": 0.1,
        "subcatchment_code": "MAP-USUTHU",
        "note": "Interim Target Instream Flow from IIMA Table 5-4."
    },
    {
        "river": "Usuthu",
        "key_point": "Big Bend (GS 16)",
        "station_code": "GS-16",
        "mean_annual_mm3": 520.0,
        "min_flow_m3_s": 1.7,
        "subcatchment_code": "MAP-USUTHU",
        "note": "Interim Target Instream Flow from IIMA Table 5-4."
    },
    {
        "river": "Mpuluzi",
        "key_point": "Dumbarton",
        "station_code": "DUMBARTON-01",
        "mean_annual_mm3": 65.0,
        "min_flow_m3_s": 0.1,
        "subcatchment_code": "MAP-MPULUZI",
        "note": "Interim Target Instream Flow from IIMA Table 5-4."
    },
    {
        "river": "Lusushwana",
        "key_point": "GS 33",
        "station_code": "GS-33",
        "mean_annual_mm3": 35.0,
        "min_flow_m3_s": 0.1,
        "subcatchment_code": "MAP-LUSUSHWANA",
        "note": "Interim Target Instream Flow from IIMA Table 5-4."
    }
]

# 3. Parse Hydrology Excel file
# Sheets: ['GS 8 ', 'GS 21', 'GS 23', 'GS 25 ', 'GS 31', 'GS 39 ']
for sheet_name in wb_hyd.sheetnames:
    ws = wb_hyd[sheet_name]
    sheet_clean = sheet_name.strip()
    
    # Map sheet name to station code
    if sheet_clean == "GS 8":
        station_code = "GS-08"
        start_row = 4
        date_col = 1
        flow_col = 3
    elif sheet_clean == "GS 21":
        station_code = "GS-21"
        start_row = 3
        date_col = 2
        flow_col = 4
    elif sheet_clean == "GS 23":
        station_code = "GS-23"
        start_row = 3
        date_col = 2
        flow_col = 4
    elif sheet_clean == "GS 25":
        station_code = "GS-25"
        start_row = 3
        date_col = 1
        flow_col = 3
    elif sheet_clean == "GS 31":
        station_code = "GS-31"
        start_row = 3
        date_col = 2
        flow_col = 4
    elif sheet_clean == "GS 39":
        station_code = "GS-39"
        start_row = 4
        date_col = 1
        flow_col = 3
    else:
        continue
        
    for r in range(start_row, ws.max_row + 1):
        date_val = parse_date(ws.cell(row=r, column=date_col).value)
        flow_val = clean_float(ws.cell(row=r, column=flow_col).value)
        if date_val and flow_val is not None:
            data["measurements"].append({
                "station_code": station_code,
                "type": "flow",
                "value": flow_val,
                "unit": "m³/s",
                "date": date_val
            })

# 4. Parse Dams Excel file
# Sheets: ['Mnjoli Dam', 'Lubovane Dam', 'Maguga Dam', 'Lavumisa Balancing Dam', 'Luphohlo Dam']
dam_mappings = {
    "Mnjoli Dam": {"code": "MNJOLI-DAM-01", "level_col": 2, "vol_col": 3, "pct_col": 4, "vol_unit": "Mm³"},
    "Lubovane Dam": {"code": "LUBOVANE-DAM-01", "level_col": 2, "vol_col": 3, "pct_col": 4, "vol_unit": "Mm³"},
    "Maguga Dam": {"code": "MAGUGA-DAM-01", "level_col": 2, "vol_col": 3, "pct_col": 4, "vol_unit": "Mm³"},
    "Lavumisa Balancing Dam": {"code": "LAVUMISA-DAM-01", "level_col": 2, "vol_col": 3, "pct_col": 4, "vol_unit": "m³"},
    "Luphohlo Dam": {"code": "LUPHOHLO-DAM-01", "level_col": 2, "vol_col": 3, "pct_col": 4, "vol_unit": "m³"}
}

for sheet_name in wb_dam.sheetnames:
    if sheet_name not in dam_mappings:
        continue
    ws = wb_dam[sheet_name]
    mapping = dam_mappings[sheet_name]
    station_code = mapping["code"]
    
    for r in range(3, ws.max_row + 1):
        date_val = parse_date(ws.cell(row=r, column=1).value)
        pct_val = clean_float(ws.cell(row=r, column=mapping["pct_col"]).value)
        vol_val = clean_float(ws.cell(row=r, column=mapping["vol_col"]).value)
        
        if date_val and pct_val is not None:
            data["measurements"].append({
                "station_code": station_code,
                "type": "dam_level",
                "value": pct_val,
                "unit": "%",
                "fsc": vol_val,  # Store volume as FSC
                "date": date_val
            })

# Save JSON file
output_path = "./database/seeders/excel_data.json"
with open(output_path, "w") as f:
    json.dump(data, f, indent=4)

print(f"Excel parsing complete! Extracted {len(data['measurements'])} measurement readings. JSON saved to {output_path}")
